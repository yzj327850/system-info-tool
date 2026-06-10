#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电脑信息获取工具 v3.0 - 终极版
内置命令实现，不依赖系统 wmic/getmac/vol/ipconfig
"""

import subprocess
import re
import json
import urllib.request
import socket
import sys
import os
import struct
import ctypes
from ctypes import wintypes
from datetime import datetime

# Windows API 常量
GENERIC_READ = 0x80000000
GENERIC_WRITE = 0x40000000
FILE_SHARE_READ = 0x00000001
FILE_SHARE_WRITE = 0x00000002
OPEN_EXISTING = 3

# 设置控制台
if sys.platform == 'win32':
    try:
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleOutputCP(65001)
        hStdOut = kernel32.GetStdHandle(-11)
        mode = wintypes.DWORD()
        kernel32.GetConsoleMode(hStdOut, ctypes.byref(mode))
        kernel32.SetConsoleMode(hStdOut, mode.value | 4)
    except:
        pass


# ==================== 内置命令实现 ====================

class HardwareInfo:
    """使用 Windows API 直接获取硬件信息，不依赖外部命令"""
    
    @staticmethod
    def get_mac_addresses():
        """使用 Windows API 获取 MAC 地址"""
        macs = []
        try:
            from ctypes import wintypes
            
            class IP_ADAPTER_INFO(ctypes.Structure):
                pass
            
            IP_ADAPTER_INFO._fields_ = [
                ("Next", ctypes.POINTER(IP_ADAPTER_INFO)),
                ("ComboIndex", wintypes.DWORD),
                ("AdapterName", ctypes.c_char * 260),
                ("Description", ctypes.c_char * 132),
                ("AddressLength", wintypes.DWORD),
                ("Address", ctypes.c_byte * 8),
                ("Index", wintypes.DWORD),
                ("Type", wintypes.DWORD),
                ("DhcpEnabled", wintypes.DWORD),
                ("CurrentIpAddress", ctypes.c_void_p),
                ("IpAddressList", ctypes.c_byte * 16),
                ("GatewayList", ctypes.c_byte * 16),
                ("DhcpServer", ctypes.c_byte * 16),
                ("HaveWins", ctypes.c_byte),
                ("PrimaryWinsServer", ctypes.c_byte * 16),
                ("SecondaryWinsServer", ctypes.c_byte * 16),
                ("LeaseObtained", wintypes.ULONG),
                ("LeaseExpires", wintypes.ULONG)
            ]
            
            GetAdaptersInfo = ctypes.windll.iphlpapi.GetAdaptersInfo
            GetAdaptersInfo.argtypes = [ctypes.POINTER(IP_ADAPTER_INFO), ctypes.POINTER(wintypes.ULONG)]
            GetAdaptersInfo.restype = wintypes.ULONG
            
            size = wintypes.ULONG(0)
            GetAdaptersInfo(None, ctypes.byref(size))
            
            if size.value > 0:
                buffer = ctypes.create_string_buffer(size.value)
                adapter = ctypes.cast(buffer, ctypes.POINTER(IP_ADAPTER_INFO))
                
                if GetAdaptersInfo(adapter, ctypes.byref(size)) == 0:
                    while adapter:
                        addr_len = adapter.contents.AddressLength
                        if addr_len == 6:
                            mac_bytes = adapter.contents.Address[:6]
                            # 将字节转换为无符号整数（c_byte是有符号的，可能导致负数）
                            mac = ':'.join(f'{(b & 0xFF):02X}' for b in mac_bytes)
                            desc = adapter.contents.Description.decode('gbk', errors='ignore').strip()
                            
                            if not any(x in desc.lower() for x in ['virtual', 'vmware', 'hyper-v', 'virtualbox', 'tunnel', 'loopback']):
                                macs.append({"name": desc, "mac": mac})
                        
                        if adapter.contents.Next:
                            adapter = adapter.contents.Next
                        else:
                            break
        except Exception as e:
            pass
        
        if not macs:
            try:
                import uuid
                node = uuid.getnode()
                if node:
                    mac = ':'.join(f'{(node >> (i * 8)) & 0xff:02X}' for i in range(5, -1, -1))
                    macs.append({"name": "主网卡", "mac": mac})
            except:
                pass
        
        return macs if macs else [{"name": "无法获取", "mac": "N/A"}]
    
    @staticmethod
    def get_cpu_info():
        """使用 Windows API 获取 CPU 信息"""
        result_parts = []
        
        # 方法1: 使用 CPUID 指令获取序列号 (Intel CPU)
        try:
            cpu_serial = HardwareInfo._get_cpu_serial_cpuid()
            if cpu_serial and cpu_serial != "N/A":
                result_parts.append(f"序列号: {cpu_serial}")
        except:
            pass
        
        # 方法2: 从注册表获取信息
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, 
                                r"HARDWARE\DESCRIPTION\System\CentralProcessor\0")
            
            processor_name = winreg.QueryValueEx(key, "ProcessorNameString")[0]
            identifier = winreg.QueryValueEx(key, "Identifier")[0]
            vendor = winreg.QueryValueEx(key, "VendorIdentifier")[0]
            
            # 尝试获取 ProcessorId (某些 CPU 支持)
            try:
                processor_id = winreg.QueryValueEx(key, "ProcessorId")[0]
                if processor_id and processor_id.strip() and not result_parts:
                    result_parts.append(f"序列号: {processor_id}")
            except:
                processor_id = None
            
            winreg.CloseKey(key)
            
            result_parts.append(f"厂商: {vendor}")
            result_parts.append(f"型号: {processor_name}")
            result_parts.append(f"标识: {identifier}")
            
        except Exception as e:
            try:
                env_id = os.environ.get('PROCESSOR_IDENTIFIER', '')
                if env_id:
                    result_parts.append(f"环境变量: {env_id}")
            except:
                pass
        
        return " | ".join(result_parts) if result_parts else "无法获取 CPU 信息"
    
    @staticmethod
    def _get_cpu_serial_cpuid():
        """使用 CPUID 指令获取 CPU 序列号"""
        try:
            # 使用 Windows API 执行 CPUID 指令
            # CPUID 指令: EAX=3 时获取处理器序列号
            
            # 定义 CPUID 结构
            class CPUID_RESULT(ctypes.Structure):
                _fields_ = [
                    ("eax", ctypes.c_uint32),
                    ("ebx", ctypes.c_uint32),
                    ("ecx", ctypes.c_uint32),
                    ("edx", ctypes.c_uint32)
                ]
            
            # 使用内联汇编或调用 cpuid
            # 这里使用 __cpuid 函数
            cpuid_result = CPUID_RESULT()
            
            # 先检查 CPU 是否支持序列号 (EAX=1, EDX bit 18)
            ctypes.windll.kernel32.GetLastError()  # 清除错误
            
            # 使用 Python 的 cpuid 替代方案 - 通过 WMI 获取
            try:
                import win32com.client
                wmi = win32com.client.Dispatch("WbemScripting.SWbemLocator")
                service = wmi.ConnectServer(".", "root\\cimv2")
                processors = service.ExecQuery("SELECT * FROM Win32_Processor")
                
                for processor in processors:
                    # ProcessorId 是唯一的处理器标识
                    processor_id = processor.ProcessorId
                    if processor_id and processor_id.strip():
                        return processor_id.strip()
            except:
                pass
            
            return "N/A"
            
        except Exception as e:
            return f"Error: {str(e)[:30]}"
    
    @staticmethod
    def get_cpu_serial_wmi():
        """使用 WMI 获取 CPU 序列号/唯一标识"""
        try:
            import win32com.client
            wmi = win32com.client.Dispatch("WbemScripting.SWbemLocator")
            service = wmi.ConnectServer(".", "root\\cimv2")
            processors = service.ExecQuery("SELECT ProcessorId, Name, Manufacturer FROM Win32_Processor")
            
            results = []
            for processor in processors:
                processor_id = processor.ProcessorId if processor.ProcessorId else "N/A"
                name = processor.Name if processor.Name else "Unknown"
                manufacturer = processor.Manufacturer if processor.Manufacturer else "Unknown"
                results.append(f"序列号: {processor_id} | {manufacturer} {name}")
            
            return "\n   ".join(results) if results else "无法获取"
            
        except Exception as e:
            return f"WMI获取失败: {str(e)[:40]}"
    
    @staticmethod
    def get_disk_info():
        """使用 Windows API 获取硬盘信息"""
        disks = []
        
        # 方法1: 使用 WMI COM 接口
        try:
            import win32com.client
            wmi = win32com.client.Dispatch("WbemScripting.SWbemLocator")
            service = wmi.ConnectServer(".", "root\\cimv2")
            
            disk_drives = service.ExecQuery("SELECT * FROM Win32_DiskDrive")
            
            for disk in disk_drives:
                try:
                    model = disk.Model if disk.Model else "Unknown"
                    serial = disk.SerialNumber if disk.SerialNumber else "N/A"
                    size_bytes = disk.Size if disk.Size else 0
                    interface = disk.InterfaceType if disk.InterfaceType else "Unknown"
                    
                    size_gb = int(size_bytes) / (1024**3) if size_bytes else 0
                    
                    disks.append({
                        "model": model,
                        "serial": serial.strip() if serial else "N/A",
                        "size": f"{size_gb:.2f} GB" if size_gb > 0 else "Unknown",
                        "interface": interface
                    })
                except:
                    continue
                    
        except Exception as e:
            # 方法2: 使用 DeviceIoControl API
            try:
                drives = HardwareInfo._get_physical_drives()
                for drive_path in drives:
                    try:
                        info = HardwareInfo._get_disk_info_api(drive_path)
                        if info:
                            disks.append(info)
                    except:
                        continue
            except:
                pass
        
        return disks if disks else [{"model": "无法获取", "serial": "N/A", "size": "N/A", "interface": "N/A"}]
    
    @staticmethod
    def _get_physical_drives():
        """获取物理磁盘路径列表"""
        drives = []
        for i in range(16):
            path = f"\\\\.\\PhysicalDrive{i}"
            try:
                handle = ctypes.windll.kernel32.CreateFileW(
                    path, 0, FILE_SHARE_READ | FILE_SHARE_WRITE,
                    None, OPEN_EXISTING, 0, None
                )
                if handle != -1:
                    drives.append(path)
                    ctypes.windll.kernel32.CloseHandle(handle)
            except:
                continue
        return drives
    
    @staticmethod
    def _get_disk_info_api(drive_path):
        """使用 DeviceIoControl 获取磁盘信息"""
        try:
            IOCTL_DISK_GET_DRIVE_GEOMETRY = 0x00070000
            
            handle = ctypes.windll.kernel32.CreateFileW(
                drive_path, GENERIC_READ | GENERIC_WRITE,
                FILE_SHARE_READ | FILE_SHARE_WRITE,
                None, OPEN_EXISTING, 0, None
            )
            
            if handle == -1:
                return None
            
            geometry = ctypes.create_string_buffer(24)
            bytes_returned = wintypes.DWORD(0)
            
            result = ctypes.windll.kernel32.DeviceIoControl(
                handle, IOCTL_DISK_GET_DRIVE_GEOMETRY,
                None, 0, geometry, 24,
                ctypes.byref(bytes_returned), None
            )
            
            size_gb = 0
            if result:
                cylinders = struct.unpack_from('<Q', geometry, 0)[0]
                tracks_per_cylinder = struct.unpack_from('<I', geometry, 12)[0]
                sectors_per_track = struct.unpack_from('<I', geometry, 16)[0]
                bytes_per_sector = struct.unpack_from('<I', geometry, 20)[0]
                
                total_bytes = cylinders * tracks_per_cylinder * sectors_per_track * bytes_per_sector
                size_gb = total_bytes / (1024**3)
            
            ctypes.windll.kernel32.CloseHandle(handle)
            
            return {
                "model": drive_path,
                "serial": "API获取",
                "size": f"{size_gb:.2f} GB" if size_gb > 0 else "Unknown",
                "interface": "Unknown"
            }
            
        except Exception as e:
            return None
    
    @staticmethod
    def get_disk_ids_diskpart():
        """使用 diskpart 获取磁盘 ID (GPT 磁盘唯一标识符)"""
        disk_ids = []
        
        try:
            import tempfile
            temp_dir = tempfile.gettempdir()
            
            # 获取磁盘列表
            temp_script = os.path.join(temp_dir, 'diskpart_list.txt')
            with open(temp_script, 'w', encoding='utf-8') as f:
                f.write("list disk\n")
            
            # 使用 GBK 编码以正确解析中文输出
            result = subprocess.run(
                f"diskpart /s \"{temp_script}\"",
                shell=True, capture_output=True, text=True,
                encoding='gbk', errors='ignore', timeout=30
            )
            
            disk_numbers = []
            for line in result.stdout.split('\n'):
                # 支持中英文 - 匹配 "磁盘 0" 或 "Disk 0" 格式
                match = re.search(r'磁盘\s+(\d+)', line, re.IGNORECASE)
                if not match:
                    match = re.search(r'Disk\s+(\d+)', line, re.IGNORECASE)
                if match:
                    num = match.group(1)
                    if num and int(num) not in disk_numbers:
                        disk_numbers.append(int(num))
            
            for disk_num in disk_numbers:
                try:
                    detail_script = os.path.join(temp_dir, f'diskpart_detail_{disk_num}.txt')
                    with open(detail_script, 'w', encoding='utf-8') as f:
                        f.write(f"select disk {disk_num}\ndetail disk\n")
                    
                    # 使用 GBK 编码以正确解析中文输出
                    result = subprocess.run(
                        f"diskpart /s \"{detail_script}\"",
                        shell=True, capture_output=True, text=True,
                        encoding='gbk', errors='ignore', timeout=30
                    )
                    
                    disk_id = "N/A"
                    disk_type = "Unknown"
                    disk_status = "Unknown"
                    
                    for line in result.stdout.split('\n'):
                        # 支持中英文磁盘ID格式
                        match = re.search(r'磁盘\s*ID[\s:]+\{?([0-9A-Fa-f\-]+)\}?', line, re.IGNORECASE)
                        if not match:
                            match = re.search(r'Disk\s*ID[\s:]+\{?([0-9A-Fa-f\-]+)\}?', line, re.IGNORECASE)
                        if not match:
                            match = re.search(r'ID[\s:]+\{([0-9A-Fa-f\-]+)\}', line, re.IGNORECASE)
                        if not match:
                            match = re.search(r'[\s:]+([0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12})', line)
                        if match:
                            disk_id = match.group(1)
                        
                        # 匹配类型
                        if "类型" in line or "Type" in line:
                            type_match = re.search(r'[:\s]+(\w+)', line)
                            if type_match:
                                disk_type = type_match.group(1)
                        
                        # 匹配状态
                        if "状态" in line or "Status" in line:
                            if "联机" in line or "Online" in line:
                                disk_status = "Online"
                            elif "脱机" in line or "Offline" in line:
                                disk_status = "Offline"
                    
                    disk_ids.append({
                        "disk_num": disk_num,
                        "disk_id": disk_id,
                        "type": disk_type,
                        "status": disk_status
                    })
                    
                except Exception as e:
                    disk_ids.append({
                        "disk_num": disk_num,
                        "disk_id": f"Error: {str(e)[:30]}",
                        "type": "Unknown",
                        "status": "Unknown"
                    })
            
            try:
                os.remove(temp_script)
            except:
                pass
                
        except Exception as e:
            disk_ids.append({
                "disk_num": -1,
                "disk_id": f"无法获取: {str(e)[:30]}",
                "type": "Unknown",
                "status": "Unknown"
            })
        
        return disk_ids if disk_ids else [{"disk_num": -1, "disk_id": "N/A", "type": "Unknown", "status": "Unknown"}]
    
    @staticmethod
    def get_volume_serials():
        """获取卷标序列号（C: D: 等）"""
        volumes = []
        
        try:
            GetVolumeInformationW = ctypes.windll.kernel32.GetVolumeInformationW
            
            for drive_letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
                drive_path = f"{drive_letter}:\\"
                
                volume_name = ctypes.create_unicode_buffer(256)
                volume_serial = wintypes.DWORD(0)
                max_component_length = wintypes.DWORD(0)
                file_system_flags = wintypes.DWORD(0)
                file_system_name = ctypes.create_unicode_buffer(256)
                
                result = GetVolumeInformationW(
                    drive_path, volume_name, 256,
                    ctypes.byref(volume_serial), ctypes.byref(max_component_length),
                    ctypes.byref(file_system_flags), file_system_name, 256
                )
                
                if result:
                    serial_hex = f"{volume_serial.value:08X}"
                    volumes.append({
                        "drive": f"{drive_letter}:",
                        "serial": f"{serial_hex[:4]}-{serial_hex[4:]}",
                        "name": volume_name.value if volume_name.value else ""
                    })
        except Exception as e:
            pass
        
        return volumes if volumes else [{"drive": "N/A", "serial": "无法获取", "name": ""}]
    
    @staticmethod
    def get_local_ips():
        """获取本地 IP 地址"""
        ips = []
        hostname = "Unknown"
        
        try:
            hostname = socket.gethostname()
        except:
            pass
        
        try:
            from ctypes import wintypes
            
            class IP_ADDR_STRING(ctypes.Structure):
                pass
            
            IP_ADDR_STRING._fields_ = [
                ("Next", ctypes.POINTER(IP_ADDR_STRING)),
                ("IpAddress", ctypes.c_char * 16),
                ("IpMask", ctypes.c_char * 16),
                ("Context", wintypes.DWORD)
            ]
            
            class IP_ADAPTER_INFO(ctypes.Structure):
                pass
            
            IP_ADAPTER_INFO._fields_ = [
                ("Next", ctypes.POINTER(IP_ADAPTER_INFO)),
                ("ComboIndex", wintypes.DWORD),
                ("AdapterName", ctypes.c_char * 260),
                ("Description", ctypes.c_char * 132),
                ("AddressLength", wintypes.DWORD),
                ("Address", ctypes.c_byte * 8),
                ("Index", wintypes.DWORD),
                ("Type", wintypes.DWORD),
                ("DhcpEnabled", wintypes.DWORD),
                ("CurrentIpAddress", ctypes.c_void_p),
                ("IpAddressList", IP_ADDR_STRING),
                ("GatewayList", IP_ADDR_STRING),
                ("DhcpServer", IP_ADDR_STRING),
                ("HaveWins", ctypes.c_byte),
                ("PrimaryWinsServer", IP_ADDR_STRING),
                ("SecondaryWinsServer", IP_ADDR_STRING),
                ("LeaseObtained", wintypes.ULONG),
                ("LeaseExpires", wintypes.ULONG)
            ]
            
            GetAdaptersInfo = ctypes.windll.iphlpapi.GetAdaptersInfo
            GetAdaptersInfo.argtypes = [ctypes.POINTER(IP_ADAPTER_INFO), ctypes.POINTER(wintypes.ULONG)]
            GetAdaptersInfo.restype = wintypes.ULONG
            
            size = wintypes.ULONG(0)
            GetAdaptersInfo(None, ctypes.byref(size))
            
            if size.value > 0:
                buffer = ctypes.create_string_buffer(size.value)
                adapter = ctypes.cast(buffer, ctypes.POINTER(IP_ADAPTER_INFO))
                
                if GetAdaptersInfo(adapter, ctypes.byref(size)) == 0:
                    while adapter:
                        ip_addr = adapter.contents.IpAddressList.IpAddress.decode('ascii', errors='ignore').strip()
                        if ip_addr and ip_addr != '0.0.0.0' and not ip_addr.startswith('127.'):
                            desc = adapter.contents.Description.decode('gbk', errors='ignore').strip()
                            if not any(x in desc.lower() for x in ['virtual', 'vmware', 'hyper-v', 'virtualbox', 'tunnel', 'loopback']):
                                ips.append({"adapter": desc, "ip": ip_addr})
                        
                        if adapter.contents.Next:
                            adapter = adapter.contents.Next
                        else:
                            break
                            
        except Exception as e:
            try:
                ip = socket.gethostbyname(hostname)
                if ip and not ip.startswith('127.'):
                    ips.append({"adapter": "默认网卡", "ip": ip})
            except:
                pass
        
        return {"hostname": hostname, "ips": ips if ips else [{"adapter": "无法获取", "ip": "N/A"}]}


# ==================== IP 归属地查询 ====================

def get_ip_from_baidu():
    """从百度智能云获取 IP 信息（精确到城市）"""
    try:
        url = "https://qifu-api.baidubce.com/ip/local/geo/v1/district"
        
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.0',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Referer': 'https://cloud.baidu.com/'
        })
        
        with urllib.request.urlopen(req, timeout=15) as response:
            data = json.loads(response.read().decode('utf-8'))
            
            if data.get('code') == 'Success':
                ip_info = data.get('data', {})
                return {
                    "ip": ip_info.get('ip', 'N/A'),
                    "country": ip_info.get('country', 'N/A'),
                    "province": ip_info.get('prov', 'N/A'),
                    "city": ip_info.get('city', 'N/A'),
                    "district": ip_info.get('district', 'N/A'),
                    "isp": ip_info.get('isp', 'N/A'),
                    "lat": ip_info.get('lat', 'N/A'),
                    "lng": ip_info.get('lng', 'N/A'),
                    "source": "百度智能云"
                }
            else:
                return {"error": f"API返回错误: {data.get('code')}"}
                
    except Exception as e:
        return {"error": f"百度API请求失败: {str(e)}"}


def get_ip_from_pconline():
    """从太平洋电脑网获取（国内精确）"""
    try:
        url = "https://whois.pconline.com.cn/ipJson.jsp?json=true"
        
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleKit/537.36',
            'Referer': 'https://whois.pconline.com.cn/'
        })
        
        with urllib.request.urlopen(req, timeout=10) as response:
            content = response.read()
            try:
                data = json.loads(content.decode('utf-8'))
            except:
                data = json.loads(content.decode('gbk', errors='ignore'))
            
            addr = data.get('addr', '')
            isp = ''
            if ' ' in addr:
                parts = addr.split(' ')
                isp = parts[-1]
                
            return {
                "ip": data.get('ip', 'N/A'),
                "country": "中国" if data.get('proCode') else 'N/A',
                "province": data.get('pro', 'N/A'),
                "city": data.get('city', 'N/A'),
                "district": "N/A",
                "isp": isp,
                "lat": "N/A",
                "lng": "N/A",
                "source": "pconline.com.cn"
            }
                
    except Exception as e:
        return {"error": f"pconline请求失败: {str(e)}"}


def get_ip_from_ipapi():
    """备用：从 ip-api 获取 IP 信息"""
    try:
        url = "http://ip-api.com/json/?lang=zh-CN&fields=status,message,country,regionName,city,isp,org,as,query"
        
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            
            if data.get('status') == 'success':
                return {
                    "ip": data.get('query', 'N/A'),
                    "country": data.get('country', 'N/A'),
                    "province": data.get('regionName', 'N/A'),
                    "city": data.get('city', 'N/A'),
                    "district": "N/A",
                    "isp": data.get('isp', 'N/A'),
                    "lat": "N/A",
                    "lng": "N/A",
                    "source": "ip-api.com"
                }
            else:
                return {"error": f"API返回错误: {data.get('message', 'Unknown')}"}
                
    except Exception as e:
        return {"error": f"ip-api请求失败: {str(e)}"}


def save_to_bidding_template(template_path, output_path, device_type, macs, cpu_serial, disks, disk_ids, ip_info):
    """按投标模板格式保存设备信息
    
    device_type: 1=报名设备, 2=报价设备, 3=上传设备
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
    except ImportError:
        raise ImportError("请先安装 openpyxl: pip install openpyxl")
    
    # 加载模板
    wb = load_workbook(template_path)
    ws = wb.active
    
    # 根据设备类型确定目标列
    if device_type == 1:  # 报名设备
        ip_col = 'I'
        location_col = 'J'
        mac_col = 'K'
        cpu_col = 'L'
        disk_serial_col = 'M'
        disk_id_col = 'N'
        device_name = "报名设备"
    elif device_type == 2:  # 报价设备
        ip_col = 'T'
        location_col = 'U'
        mac_col = 'V'
        cpu_col = 'W'
        disk_serial_col = 'X'
        disk_id_col = 'Y'
        device_name = "报价设备"
    elif device_type == 3:  # 上传设备
        ip_col = 'AE'
        location_col = 'AF'
        mac_col = 'AG'
        cpu_col = 'AH'
        disk_serial_col = 'AI'
        disk_id_col = 'AJ'
        device_name = "上传设备"
    else:
        raise ValueError(f"无效的设备类型: {device_type}")
    
    row = 4  # 数据从第4行开始
    
    # 设置单元格对齐方式（自动换行）
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # 1. 公网IP
    ip_cell = ws[f'{ip_col}{row}']
    ip_cell.value = ip_info.get('ip', 'N/A')
    ip_cell.alignment = wrap_alignment
    
    # 2. IP归属地 (格式: XX省XX市)
    province = ip_info.get('province', '')
    city = ip_info.get('city', '')
    if province and city and province != 'N/A' and city != 'N/A':
        location = f"{province}{city}"
    else:
        location = ip_info.get('country', 'N/A')
    location_cell = ws[f'{location_col}{row}']
    location_cell.value = location
    location_cell.alignment = wrap_alignment
    
    # 3. MAC地址 (多个分行显示，只保留MAC，不保留网卡名)
    mac_list = [m['mac'] for m in macs if m['mac'] != 'N/A' and 'virtual' not in m['name'].lower()]
    mac_value = '\n'.join(mac_list) if mac_list else 'N/A'
    mac_cell = ws[f'{mac_col}{row}']
    mac_cell.value = mac_value
    mac_cell.alignment = wrap_alignment
    
    # 4. CPU序列号 (只保留序列号)
    cpu_value = cpu_serial if cpu_serial and cpu_serial != 'N/A' else 'N/A'
    # 如果包含其他信息，尝试提取序列号
    if '|' in cpu_value:
        parts = cpu_value.split('|')
        for part in parts:
            if '序列号' in part or 'ProcessorId' in part:
                cpu_value = part.split(':')[-1].strip()
                break
    cpu_cell = ws[f'{cpu_col}{row}']
    cpu_cell.value = cpu_value
    cpu_cell.alignment = wrap_alignment
    
    # 5. 硬盘序列号 (多个分行显示，只保留序列号)
    disk_serial_list = []
    for disk in disks:
        serial = disk.get('serial', 'N/A')
        if serial and serial != 'N/A':
            disk_serial_list.append(serial)
    disk_serial_value = '\n'.join(disk_serial_list) if disk_serial_list else 'N/A'
    disk_serial_cell = ws[f'{disk_serial_col}{row}']
    disk_serial_cell.value = disk_serial_value
    disk_serial_cell.alignment = wrap_alignment
    
    # 6. 磁盘ID (多个分行显示，只保留ID)
    disk_id_list = []
    for disk in disk_ids:
        disk_id = disk.get('disk_id', 'N/A')
        if disk_id and disk_id != 'N/A' and not disk_id.startswith('无法获取') and not disk_id.startswith('Error'):
            disk_id_list.append(disk_id)
    disk_id_value = '\n'.join(disk_id_list) if disk_id_list else 'N/A'
    disk_id_cell = ws[f'{disk_id_col}{row}']
    disk_id_cell.value = disk_id_value
    disk_id_cell.alignment = wrap_alignment
    
    # 保存文件
    wb.save(output_path)
    return device_name


def save_to_excel(filename, macs, cpu_info, cpu_serial, disks, disk_ids, volumes, local_info, ip_info):
    """保存信息到 Excel 文件（原始格式）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("请先安装 openpyxl: pip install openpyxl")
    
    wb = Workbook()
    
    # 定义样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== Sheet 1: 概览 ==========
    ws1 = wb.active
    ws1.title = "概览"
    
    # 标题
    ws1['A1'] = "电脑信息获取报告"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal="center")
    ws1.merge_cells('A1:D1')
    
    ws1['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1['A2'].alignment = Alignment(horizontal="center")
    ws1.merge_cells('A2:D2')
    
    ws1['A3'] = f"主机名: {local_info['hostname']}"
    ws1['A3'].alignment = Alignment(horizontal="center")
    ws1.merge_cells('A3:D3')
    
    row = 5
    
    # 公网 IP 信息
    ws1[f'A{row}'] = "公网 IP 信息"
    ws1[f'A{row}'].font = Font(bold=True, size=12)
    ws1.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ip_data = [
        ["项目", "值"],
        ["公网 IP", ip_info.get('ip', 'N/A')],
        ["国家/地区", ip_info.get('country', 'N/A')],
        ["省份", ip_info.get('province', 'N/A')],
        ["城市", ip_info.get('city', 'N/A')],
        ["区县", ip_info.get('district', 'N/A')],
        ["运营商", ip_info.get('isp', 'N/A')],
        ["数据来源", ip_info.get('source', 'N/A')],
    ]
    
    for r_idx, row_data in enumerate(ip_data, start=row):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = cell_alignment
            if r_idx == row:  # 表头
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    row += len(ip_data) + 2
    
    # CPU 信息
    ws1[f'A{row}'] = "CPU 信息"
    ws1[f'A{row}'].font = Font(bold=True, size=12)
    ws1.merge_cells(f'A{row}:D{row}')
    row += 1
    
    cpu_data = [
        ["项目", "值"],
        ["CPU 详情", cpu_info],
        ["CPU 序列号", cpu_serial],
    ]
    
    for r_idx, row_data in enumerate(cpu_data, start=row):
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = cell_alignment
            if r_idx == row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    # 调整列宽
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 60
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20
    
    # ========== Sheet 2: MAC 地址 ==========
    ws2 = wb.create_sheet("MAC 地址")
    headers = ["网卡名称", "MAC 地址"]
    for c_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for r_idx, mac_info in enumerate(macs, start=2):
        ws2.cell(row=r_idx, column=1, value=mac_info['name']).border = thin_border
        ws2.cell(row=r_idx, column=2, value=mac_info['mac']).border = thin_border
    
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 25
    
    # ========== Sheet 3: 硬盘信息 ==========
    ws3 = wb.create_sheet("硬盘信息")
    headers = ["型号", "序列号", "容量", "接口类型"]
    for c_idx, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for r_idx, disk in enumerate(disks, start=2):
        ws3.cell(row=r_idx, column=1, value=disk['model']).border = thin_border
        ws3.cell(row=r_idx, column=2, value=disk['serial']).border = thin_border
        ws3.cell(row=r_idx, column=3, value=disk['size']).border = thin_border
        ws3.cell(row=r_idx, column=4, value=disk['interface']).border = thin_border
    
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 15
    
    # ========== Sheet 4: 磁盘 ID ==========
    ws4 = wb.create_sheet("磁盘 ID")
    headers = ["磁盘号", "磁盘 ID", "类型", "状态"]
    for c_idx, header in enumerate(headers, start=1):
        cell = ws4.cell(row=1, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for r_idx, disk in enumerate(disk_ids, start=2):
        ws4.cell(row=r_idx, column=1, value=f"磁盘 {disk['disk_num']}").border = thin_border
        ws4.cell(row=r_idx, column=2, value=disk['disk_id']).border = thin_border
        ws4.cell(row=r_idx, column=3, value=disk['type']).border = thin_border
        ws4.cell(row=r_idx, column=4, value=disk['status']).border = thin_border
    
    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 45
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 15
    
    # ========== Sheet 5: 卷标序列号 ==========
    ws5 = wb.create_sheet("卷标序列号")
    headers = ["分区", "序列号", "卷标名"]
    for c_idx, header in enumerate(headers, start=1):
        cell = ws5.cell(row=1, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for r_idx, vol in enumerate(volumes, start=2):
        ws5.cell(row=r_idx, column=1, value=vol['drive']).border = thin_border
        ws5.cell(row=r_idx, column=2, value=vol['serial']).border = thin_border
        ws5.cell(row=r_idx, column=3, value=vol['name']).border = thin_border
    
    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 20
    ws5.column_dimensions['C'].width = 25
    
    # ========== Sheet 6: IP 地址 ==========
    ws6 = wb.create_sheet("IP 地址")
    
    # 本地 IP
    ws6['A1'] = "本地 IP 地址"
    ws6['A1'].font = Font(bold=True, size=12)
    ws6.merge_cells('A1:B1')
    
    headers = ["网卡", "IP 地址"]
    for c_idx, header in enumerate(headers, start=1):
        cell = ws6.cell(row=2, column=c_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for r_idx, ip in enumerate(local_info['ips'], start=3):
        ws6.cell(row=r_idx, column=1, value=ip['adapter']).border = thin_border
        ws6.cell(row=r_idx, column=2, value=ip['ip']).border = thin_border
    
    row = len(local_info['ips']) + 5
    
    # 公网 IP
    ws6[f'A{row}'] = "公网 IP 信息"
    ws6[f'A{row}'].font = Font(bold=True, size=12)
    ws6.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ip_details = [
        ["公网 IP", ip_info.get('ip', 'N/A')],
        ["国家/地区", ip_info.get('country', 'N/A')],
        ["省份", ip_info.get('province', 'N/A')],
        ["城市", ip_info.get('city', 'N/A')],
        ["区县", ip_info.get('district', 'N/A')],
        ["运营商", ip_info.get('isp', 'N/A')],
        ["数据来源", ip_info.get('source', 'N/A')],
    ]
    
    for r_idx, (key, value) in enumerate(ip_details, start=row):
        ws6.cell(row=r_idx, column=1, value=key).border = thin_border
        ws6.cell(row=r_idx, column=2, value=value).border = thin_border
    
    ws6.column_dimensions['A'].width = 50
    ws6.column_dimensions['B'].width = 30
    
    # 保存文件
    wb.save(filename)


def get_ip_info():
    """获取 IP 地址及归属地 - 多源获取，精确到城市"""
    sources = [
        ("百度智能云", get_ip_from_baidu),
        ("太平洋电脑网", get_ip_from_pconline),
        ("ip-api.com", get_ip_from_ipapi)
    ]
    
    errors = []
    
    for name, func in sources:
        try:
            result = func()
            if "error" not in result:
                return result
            else:
                errors.append(f"{name}: {result['error']}")
        except Exception as e:
            errors.append(f"{name}: {str(e)}")
    
    try:
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
        return {
            "ip": local_ip,
            "country": "本地网络",
            "province": "N/A",
            "city": "N/A",
            "district": "N/A",
            "isp": "N/A",
            "lat": "N/A",
            "lng": "N/A",
            "source": "本地获取 (公网查询失败)",
            "errors": "; ".join(errors[:2])
        }
    except:
        return {
            "ip": "N/A",
            "country": "N/A",
            "province": "N/A", 
            "city": "N/A",
            "district": "N/A",
            "isp": "N/A",
            "lat": "N/A",
            "lng": "N/A",
            "source": "所有方法失败",
            "errors": "; ".join(errors[:2])
        }


# ==================== 主程序 ====================

def main():
    """主函数"""
    print("=" * 70)
    print("                    电脑信息获取工具 v3.0")
    print("              (内置API实现，无需系统命令)")
    print("=" * 70)
    print()
    
    hw = HardwareInfo()
    
    # 1. MAC 地址
    print("【1. MAC 地址】")
    macs = hw.get_mac_addresses()
    for mac_info in macs:
        print(f"   网卡: {mac_info['name']}")
        print(f"   MAC:  {mac_info['mac']}")
        print()
    
    # 2. CPU 信息
    print("【2. CPU 信息】")
    cpu_info = hw.get_cpu_info()
    print(f"   {cpu_info}")
    print()
    
    # 2.1 CPU 序列号 (单独获取)
    print("【2.1 CPU 序列号 (WMI)】")
    cpu_serial = hw.get_cpu_serial_wmi()
    print(f"   {cpu_serial}")
    print()
    
    # 3. 硬盘序列号
    print("【3. 硬盘信息】")
    disks = hw.get_disk_info()
    for disk in disks:
        print(f"   型号: {disk['model']}")
        print(f"   序列号: {disk['serial']}")
        print(f"   容量: {disk['size']}")
        print(f"   接口: {disk['interface']}")
        print()
    
    # 4. 磁盘 ID (diskpart 方式 - GPT磁盘ID)
    print("【4. 磁盘 ID (diskpart GPT标识符)】")
    print("   正在使用 diskpart 获取磁盘 ID...")
    disk_ids = hw.get_disk_ids_diskpart()
    for disk in disk_ids:
        print(f"   磁盘 {disk['disk_num']}: {disk['disk_id']}")
        print(f"            类型: {disk['type']} | 状态: {disk['status']}")
    print()
    
    # 5. 卷标序列号 (C: D: 等)
    print("【5. 卷标序列号 (分区卷标)】")
    volumes = hw.get_volume_serials()
    for vol in volumes:
        print(f"   分区 {vol['drive']} -> 序列号: {vol['serial']}")
        if vol['name']:
            print(f"            卷标名: {vol['name']}")
    print()
    
    # 6. IP 地址及归属地
    print("【6. IP 地址及归属地】")
    
    # 本地 IP
    local_info = hw.get_local_ips()
    print(f"   主机名: {local_info['hostname']}")
    print(f"   本地 IP:")
    for ip_info in local_info['ips']:
        print(f"      {ip_info['adapter']}: {ip_info['ip']}")
    print()
    
    # 公网 IP 及归属地
    print("   正在查询公网 IP 信息...")
    ip_info = get_ip_info()
    print(f"   公网 IP: {ip_info.get('ip', 'N/A')}")
    print(f"   国家/地区: {ip_info.get('country', 'N/A')}")
    print(f"   省份: {ip_info.get('province', 'N/A')}")
    print(f"   城市: {ip_info.get('city', 'N/A')}")
    print(f"   区县: {ip_info.get('district', 'N/A')}")
    print(f"   运营商: {ip_info.get('isp', 'N/A')}")
    if ip_info.get('lat') != 'N/A':
        print(f"   纬度: {ip_info.get('lat')}")
    if ip_info.get('lng') != 'N/A':
        print(f"   经度: {ip_info.get('lng')}")
    print(f"   数据来源: {ip_info.get('source', 'N/A')}")
    if 'errors' in ip_info:
        print(f"   错误信息: {ip_info['errors']}")
    
    print()
    print("=" * 70)
    print("                    信息获取完成")
    print("=" * 70)
    
    # 保存到文件
    try:
        save_to_file = input("\n是否保存到文件? (y/n): ").strip().lower()
        if save_to_file == 'y':
            # 询问设备阶段
            print("\n" + "=" * 70)
            print("请选择设备所属阶段:")
            print("  1. 报名设备")
            print("  2. 报价设备")
            print("  3. 上传设备")
            print("=" * 70)
            
            device_type = 0
            device_name = ""
            while device_type not in [1, 2, 3]:
                try:
                    choice = input("请输入对应数字 (1/2/3): ").strip()
                    device_type = int(choice)
                    if device_type == 1:
                        device_name = "报名设备"
                    elif device_type == 2:
                        device_name = "报价设备"
                    elif device_type == 3:
                        device_name = "上传设备"
                    else:
                        print("无效选择，请重新输入")
                        device_type = 0
                except ValueError:
                    print("请输入数字 1、2 或 3")
            
            # 更新文件名，包含设备阶段信息
            txt_filename = f"电脑信息_{device_name}_{local_info['hostname']}.txt"
            
            # 保存为 TXT（包含设备阶段信息）
            with open(txt_filename, 'w', encoding='utf-8') as f:
                f.write("=" * 70 + "\n")
                f.write("                    电脑信息获取报告 v3.0\n")
                f.write("=" * 70 + "\n")
                f.write(f"设备阶段: {device_name}\n")
                f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 70 + "\n\n")
                
                f.write("【1. MAC 地址】\n")
                for mac_info in macs:
                    f.write(f"   网卡: {mac_info['name']}\n")
                    f.write(f"   MAC:  {mac_info['mac']}\n\n")
                
                f.write("【2. CPU 信息】\n")
                f.write(f"   {cpu_info}\n\n")
                
                f.write("【3. 硬盘信息】\n")
                for disk in disks:
                    f.write(f"   型号: {disk['model']}\n")
                    f.write(f"   序列号: {disk['serial']}\n")
                    f.write(f"   容量: {disk['size']}\n")
                    f.write(f"   接口: {disk['interface']}\n\n")
                
                f.write("【4. 磁盘 ID (diskpart GPT标识符)】\n")
                for disk in disk_ids:
                    f.write(f"   磁盘 {disk['disk_num']}: {disk['disk_id']}\n")
                    f.write(f"            类型: {disk['type']} | 状态: {disk['status']}\n")
                f.write("\n")
                
                f.write("【5. 卷标序列号 (分区卷标)】\n")
                for vol in volumes:
                    f.write(f"   分区 {vol['drive']} -> 序列号: {vol['serial']}\n")
                    if vol['name']:
                        f.write(f"            卷标名: {vol['name']}\n")
                f.write("\n")
                
                f.write("【6. IP 地址及归属地】\n")
                f.write(f"   主机名: {local_info['hostname']}\n")
                f.write(f"   本地 IP:\n")
                for ip in local_info['ips']:
                    f.write(f"      {ip['adapter']}: {ip['ip']}\n")
                f.write(f"\n")
                f.write(f"   公网 IP: {ip_info.get('ip', 'N/A')}\n")
                f.write(f"   国家/地区: {ip_info.get('country', 'N/A')}\n")
                f.write(f"   省份: {ip_info.get('province', 'N/A')}\n")
                f.write(f"   城市: {ip_info.get('city', 'N/A')}\n")
                f.write(f"   区县: {ip_info.get('district', 'N/A')}\n")
                f.write(f"   运营商: {ip_info.get('isp', 'N/A')}\n")
                f.write(f"   数据来源: {ip_info.get('source', 'N/A')}\n")
                if 'errors' in ip_info:
                    f.write(f"   错误信息: {ip_info['errors']}\n")
                
                f.write("\n" + "=" * 70 + "\n")
            
            print(f"\nTXT 文件已保存: {txt_filename}")
            
            # 填充投标模板
            try:
                # 获取模板文件路径（支持打包后的exe）
                if getattr(sys, 'frozen', False):
                    # 运行在打包后的exe中
                    bundle_dir = sys._MEIPASS
                else:
                    # 运行在普通Python环境中
                    bundle_dir = os.path.dirname(os.path.abspath(__file__))
                
                template_path = os.path.join(bundle_dir, 'template.xlsx')
                
                if os.path.exists(template_path):
                    bidding_filename = f"投标设备信息_{device_name}_{local_info['hostname']}.xlsx"
                    save_to_bidding_template(template_path, bidding_filename, device_type, macs, cpu_serial, disks, disk_ids, ip_info)
                    print(f"投标模板已填充并保存: {bidding_filename}")
                else:
                    print(f"\n注意: 未找到嵌入的模板文件")
            except Exception as e:
                print(f"填充投标模板时出错: {e}")
    except EOFError:
        pass
    except Exception as e:
        print(f"\n保存文件时出错: {e}")
    
    try:
        input("\n按回车键退出...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
